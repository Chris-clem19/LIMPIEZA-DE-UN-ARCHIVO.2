from openpyxl import load_workbook
import pandas as pd


## Cargar el Excel con pandas y de forma más eficiente
'''
archivo_excel = "base_datos_sucia.xlsx"
df = pd.read_excel(archivo_excel)
df.head()
'''

#Cargar archivo del Excel con openpyxl
archivo_excel = "base_datos_sucia.xlsx"
wb = load_workbook(filename=archivo_excel)
hoja = wb.active

#Extraer encabezados
encabezados = [celda.value for celda in hoja[1]]

#Extraer datos
datos_extraidos = []
for fila in hoja.iter_rows(min_row=2, values_only=True):
    datos_extraidos.append(fila)

#Crear DataFrame con pandas
df = pd.DataFrame(datos_extraidos, columns=encabezados)

#Normalizar nombres y ciudades
df["Nombre"] = df["Nombre"].astype(str).str.strip().str.title().fillna("Desconocido")
df["Ciudad"] = df["Ciudad"].astype(str).str.strip().str.lower().replace({
    "cdmx": "Ciudad de México",
    "cdm": "Ciudad de México",
    "cd": "Ciudad de México",
    "ciudad": "Ciudad de México",
    "gdl": "Guadalajara",
    "gto": "Guanajuato",
})

#Convertir edades en palabras a números
cambiar_edades = {
    "veinte": 20, "veintiuno": 21, "veintidos": 22, "veintitrés": 23, "veinticuatro": 24,
    "veinticinco": 25, "treinta": 30, "cuarenta": 40
}


## Función un poco más eficiente

def convertir_edad_v2(edad):
    try:
        return int(edad) if isinstance(edad, str) else None
    except:
        return None


def convertir_edad(edad):
    if isinstance(edad, str):
        edad = edad.strip().lower()
        return cambiar_edades.get(edad, None)
    try:
        return int(edad)
    except:
        return None

df["Edad"] = df["Edad"].apply(convertir_edad)

#Normalizar fechas
df["Fecha de Registro"] = pd.to_datetime(df["Fecha de Registro"], errors="coerce")

#Corregir errores comunes en correos
def corregir_correo(correo):
    if not isinstance(correo, str):
        return None
    correo = correo.strip().replace("@@", "@")
    if "@" in correo and "." not in correo and correo.endswith("@gmail"):
        correo += ".com"
    return correo if "@" in correo else None

df["Correo electrónico"] = df["Correo electrónico"].apply(corregir_correo)

# Reemplazar valores vacíos en campos clave
df["Nombre"] = df["Nombre"].replace("nan", "Desconocido")
df["Correo electrónico"] = df["Correo electrónico"].fillna("CORREO NO DISPONIBLE")

# Eliminar solo si faltan los dos campos más importantes
df = df.dropna(subset=["Nombre", "Correo electrónico"], how="all")

# Guardar archivo limpio
df.to_excel("BASE DE DATOS LIMPIA.xlsx", index=False)
